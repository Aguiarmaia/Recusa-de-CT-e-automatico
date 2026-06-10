# -*- coding: utf-8 -*-

import re
import os
import time
import logging
import sys
from pathlib import Path
from dataclasses import dataclass, field
from typing import Optional

import undetected_chromedriver as uc
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, WebDriverException
from urllib3.exceptions import ReadTimeoutError
from openpyxl import load_workbook

# ============================================================
# CONFIGURAÇÕES — edite apenas esta seção
# ============================================================

# Caminho do arquivo Excel (deixe vazio "" para buscar automaticamente na pasta do script)
ARQUIVO_EXCEL = ""

# Colunas (1 = A, 2 = B, etc.)
COLUNA_CNPJ      = 1   # A
COLUNA_CHAVE     = 2   # B
COLUNA_TEXTO     = 3   # C
COLUNA_STATUS    = 4   # D
COLUNA_PROTOCOLO = 5   # E

# Linha onde os dados começam (pula cabeçalho)
LINHA_INICIAL = 3

# Limites de execução
LIMITE_POR_EXECUCAO   = 500   # Máximo de linhas processadas por rodada
LIMITE_RESTART_DRIVER = 80    # Reinicia o Chrome a cada N linhas (evita travamento)
TENTATIVAS_POR_LINHA  = 3     # Tentativas antes de marcar como ERRO

# Salvar Excel a cada N linhas (0 = só no final)
SALVAR_A_CADA = 10

# Timeout do Selenium em segundos
TIMEOUT = 120

# URL do portal
URL = "https://dfe-portal.svrs.rs.gov.br/CteSSL/PrestacaoServicoDesacordo"

# ============================================================
# CONFIGURAÇÃO DE LOG
# ============================================================

def configurar_log() -> logging.Logger:
    """Configura log simultâneo no terminal e em arquivo."""
    logger = logging.getLogger("automacao_cte")
    logger.setLevel(logging.DEBUG)

    fmt = logging.Formatter("%(asctime)s [%(levelname)s] %(message)s", datefmt="%H:%M:%S")

    # Terminal
    ch = logging.StreamHandler(sys.stdout)
    ch.setLevel(logging.INFO)
    ch.setFormatter(fmt)
    logger.addHandler(ch)

    # Arquivo
    log_path = Path(__file__).parent / "automacao_cte.log"
    fh = logging.FileHandler(log_path, encoding="utf-8")
    fh.setLevel(logging.DEBUG)
    fh.setFormatter(fmt)
    logger.addHandler(fh)

    return logger

log = configurar_log()

# ============================================================
# ESTRUTURA DE DADOS
# ============================================================

@dataclass
class LinhaExcel:
    numero: int
    cnpj: str
    chave: str
    texto: str

@dataclass
class Resultado:
    total:  int = 0
    sucesso: int = 0
    erro:   int = 0
    pulado: int = 0
    erros_detalhes: list = field(default_factory=list)

# ============================================================
# EXCEL
# ============================================================

def localizar_excel() -> Optional[Path]:
    """Retorna o caminho do Excel configurado ou busca na pasta do script."""
    if ARQUIVO_EXCEL:
        p = Path(ARQUIVO_EXCEL)
        if p.exists():
            log.info(f"Excel encontrado (configurado): {p}")
            return p
        else:
            log.error(f"Arquivo configurado não encontrado: {p}")
            return None

    pasta = Path(__file__).parent
    for arquivo in pasta.iterdir():
        if arquivo.suffix.lower() in (".xlsx", ".xlsm") and not arquivo.name.startswith("~$"):
            log.info(f"Excel encontrado automaticamente: {arquivo}")
            return arquivo

    log.error("Nenhum arquivo Excel encontrado na pasta do script.")
    return None


def abrir_excel(caminho: Path):
    """Abre o workbook mantendo fórmulas calculadas."""
    try:
        # keep_links=False evita erro em arquivos com links externos
        wb = load_workbook(caminho, data_only=True, keep_links=False)
        log.info("Excel aberto com sucesso.")
        return wb, wb.active
    except Exception as e:
        log.error(f"Erro ao abrir Excel: {e}")
        return None, None


def salvar_excel(wb, caminho: Path):
    """Salva o workbook com tratamento de erro."""
    try:
        wb.save(caminho)
        log.debug("Excel salvo.")
    except PermissionError:
        log.warning("Não foi possível salvar — feche o arquivo Excel e tente novamente.")
    except Exception as e:
        log.error(f"Erro ao salvar Excel: {e}")


def ler_linhas(ws, resultado: Resultado) -> list[LinhaExcel]:
    """Lê e valida todas as linhas pendentes do Excel."""
    linhas = []

    for num in range(LINHA_INICIAL, ws.max_row + 1):
        chave     = ws.cell(row=num, column=COLUNA_CHAVE).value
        protocolo = ws.cell(row=num, column=COLUNA_PROTOCOLO).value
        cnpj      = ws.cell(row=num, column=COLUNA_CNPJ).value
        texto     = ws.cell(row=num, column=COLUNA_TEXTO).value

        # Pula linha vazia ou já processada
        if not chave:
            continue
        if protocolo:
            resultado.pulado += 1
            continue

        cnpj_limpo  = re.sub(r"\D", "", str(cnpj or ""))
        chave_limpa = re.sub(r"\D", "", str(chave or ""))
        texto_str   = str(texto or "").strip()

        if not cnpj_limpo or not chave_limpa or not texto_str:
            log.warning(f"Linha {num} ignorada — CNPJ, chave ou texto vazio após limpeza.")
            resultado.pulado += 1
            continue

        linhas.append(LinhaExcel(numero=num, cnpj=cnpj_limpo, chave=chave_limpa, texto=texto_str))

    return linhas[:LIMITE_POR_EXECUCAO]

# ============================================================
# SELENIUM / CHROME
# ============================================================

def iniciar_driver() -> uc.Chrome:
    """Inicializa o Chrome com undetected_chromedriver."""
    options = uc.ChromeOptions()
    options.add_argument("--start-maximized")
    options.add_argument("--disable-notifications")
    options.add_argument("--lang=pt-BR")
    options.add_argument("--disable-gpu")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--disable-blink-features=AutomationControlled")

    driver = uc.Chrome(options=options)
    log.debug("Driver Chrome iniciado.")
    return driver


def aguardar_autenticacao(driver: uc.Chrome, wait: WebDriverWait):
    """Abre o portal e aguarda o usuário selecionar o certificado digital."""
    driver.get(URL)
    log.info("🔐 Selecione o certificado digital no navegador...")
    wait.until(EC.presence_of_element_located((By.ID, "CodInscrMfEvento")))
    log.info("✅ Autenticado com sucesso.")


def reiniciar_driver(driver: uc.Chrome) -> tuple[uc.Chrome, WebDriverWait]:
    """Encerra e reinicia o driver, reabrindo o portal."""
    log.info("🔄 Reiniciando Chrome...")
    try:
        driver.quit()
    except Exception:
        pass
    time.sleep(2)

    driver = iniciar_driver()
    wait   = WebDriverWait(driver, TIMEOUT)
    driver.get(URL)
    wait.until(EC.presence_of_element_located((By.ID, "CodInscrMfEvento")))
    log.info("✅ Chrome reiniciado.")
    return driver, wait


def preencher_formulario(driver: uc.Chrome, wait: WebDriverWait, linha: LinhaExcel):
    """Preenche os campos do formulário."""
    campo_cnpj = wait.until(EC.presence_of_element_located((By.ID, "CodInscrMfEvento")))
    campo_cnpj.clear()
    campo_cnpj.send_keys(linha.cnpj)

    campo_chave = driver.find_element(By.ID, "ChaveAcessoDfe")
    campo_chave.clear()
    campo_chave.send_keys(linha.chave)

    # Usa CSS selector mais robusto que TAG_NAME genérico
    textarea = driver.find_element(By.CSS_SELECTOR, "textarea[id], textarea[name]")
    textarea.clear()
    textarea.send_keys(linha.texto)


def clicar_registrar(driver: uc.Chrome, wait: WebDriverWait) -> bool:
    """Clica no botão Registrar e aguarda resposta."""
    try:
        botao = wait.until(
            EC.element_to_be_clickable((By.XPATH, "//button[contains(., 'Registra')]"))
        )
        driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", botao)
        time.sleep(0.8)
        try:
            botao.click()
        except Exception:
            botao.send_keys("\n")
        time.sleep(4)
        return True
    except Exception as e:
        log.warning(f"Erro ao clicar em Registrar: {e}")
        return False


def capturar_protocolo(driver: uc.Chrome) -> str:
    """Extrai o número de protocolo da página de confirmação."""
    try:
        elemento = driver.find_element(By.XPATH, "//*[contains(text(), 'Protocolo')]")
        match = re.search(r"\d{10,}", elemento.text)
        return match.group(0) if match else ""
    except Exception as e:
        log.debug(f"Protocolo não encontrado: {e}")
        return ""

# ============================================================
# PROCESSAMENTO PRINCIPAL
# ============================================================

def processar_linha(
    driver: uc.Chrome,
    wait: WebDriverWait,
    linha: LinhaExcel,
    ws,
    wb,
    caminho: Path,
    resultado: Resultado,
) -> tuple[uc.Chrome, WebDriverWait]:
    """Processa uma linha com retry automático."""

    for tentativa in range(1, TENTATIVAS_POR_LINHA + 1):
        try:
            log.info(f"📄 Linha {linha.numero} | Tentativa {tentativa}/{TENTATIVAS_POR_LINHA}")
            preencher_formulario(driver, wait, linha)

            if not clicar_registrar(driver, wait):
                raise WebDriverException("Botão Registrar não respondeu.")

            wait.until(
                EC.presence_of_element_located(
                    (By.XPATH, "//*[contains(text(), 'Evento enviado com sucesso')]")
                )
            )

            protocolo = capturar_protocolo(driver)

            if protocolo:
                ws.cell(row=linha.numero, column=COLUNA_STATUS).value    = "OK"
                ws.cell(row=linha.numero, column=COLUNA_PROTOCOLO).value = protocolo
                resultado.sucesso += 1
                log.info(f"✅ Sucesso | Protocolo: {protocolo}")
            else:
                ws.cell(row=linha.numero, column=COLUNA_STATUS).value = "SEM_PROTOCOLO"
                resultado.erro += 1
                log.warning(f"⚠️ Evento enviado mas protocolo não capturado.")

            # Volta à página inicial para o próximo registro
            driver.get(URL)
            time.sleep(1.5)
            return driver, wait  # Sucesso — sai do loop de tentativas

        except (TimeoutException, WebDriverException, ReadTimeoutError) as e:
            log.warning(f"⚠️ Erro na tentativa {tentativa}: {type(e).__name__}")
            if tentativa < TENTATIVAS_POR_LINHA:
                log.info("🔄 Reiniciando driver antes de tentar novamente...")
                driver, wait = reiniciar_driver(driver)
            else:
                ws.cell(row=linha.numero, column=COLUNA_STATUS).value = "ERRO"
                resultado.erro += 1
                resultado.erros_detalhes.append(f"Linha {linha.numero}: {type(e).__name__}")
                log.error(f"❌ Linha {linha.numero} marcada como ERRO após {TENTATIVAS_POR_LINHA} tentativas.")
                driver, wait = reiniciar_driver(driver)

    return driver, wait


def processar_excel():
    """Função principal — orquestra todo o fluxo."""
    log.info("=" * 55)
    log.info("  Automação CTe — Prestação de Serviço em Desacordo")
    log.info("=" * 55)

    # 1. Localiza e abre o Excel
    caminho = localizar_excel()
    if not caminho:
        return

    wb, ws = abrir_excel(caminho)
    if not wb:
        return

    # 2. Lê linhas pendentes
    resultado = Resultado()
    linhas = ler_linhas(ws, resultado)

    if not linhas:
        log.info("Nenhuma linha pendente encontrada. Encerrando.")
        wb.close()
        return

    log.info(f"📋 {len(linhas)} linha(s) para processar | {resultado.pulado} já processadas (puladas).")

    # 3. Inicia o Chrome
    driver = iniciar_driver()
    wait   = WebDriverWait(driver, TIMEOUT)

    try:
        aguardar_autenticacao(driver, wait)

        contador_restart = 0

        for i, linha in enumerate(linhas):
            resultado.total += 1
            contador_restart += 1

            # Reinício preventivo a cada N linhas
            if contador_restart >= LIMITE_RESTART_DRIVER:
                driver, wait = reiniciar_driver(driver)
                contador_restart = 0

            driver, wait = processar_linha(driver, wait, linha, ws, wb, caminho, resultado)

            # Salva periodicamente
            if SALVAR_A_CADA > 0 and resultado.total % SALVAR_A_CADA == 0:
                salvar_excel(wb, caminho)

    except KeyboardInterrupt:
        log.warning("⛔ Interrompido pelo usuário (Ctrl+C).")



# ============================================================
# ENTRY POINT
# ============================================================

if __name__ == "__main__":
    processar_excel()